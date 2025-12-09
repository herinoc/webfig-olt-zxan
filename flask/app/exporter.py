import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def export_onu_history_to_excel(db):
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM config_onu_history ORDER BY created_at DESC")
    rows = cursor.fetchall()
    cursor.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "ONU History"

    # Header
    headers = ["Nama", "Alamat", "SN", "Port", "ONU", "VLAN",
               "Upload", "Download", "Username", "Password",
               "LAN", "Date"]
    ws.append(headers)

    # Style untuk header
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # hijau
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        values = [
            row.get("nama_pelanggan", ""),
            row.get("alamat", ""),
            row.get("sn", ""),
            row.get("port_base", ""),
            row.get("onu_num", ""),
            row.get("vlan", ""),
            row.get("upload_profile", ""),
            row.get("download_profile", ""),
            row.get("pppoe_username", ""),
            row.get("pppoe_password", ""),
            row.get("lan_lock", ""),
            row.get("created_at", ""),
        ]
        ws.append(values)
        for c_idx, _ in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto adjust column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
